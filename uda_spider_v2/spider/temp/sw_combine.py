import xlrd
import xlwt
import openpyxl
import string
import sys
import os
import shutil

JOB_LIST_DIR = './data/list/'
JOB_DETAIL_DIR = './data/detail/'

Pos_Table1 = JOB_DETAIL_DIR + sys.argv[1]
Pos_Table2 = JOB_LIST_DIR + sys.argv[2]
Pos_Table_End = sys.argv[2]
Count1 = int(sys.argv[3])-1
Count2 = int(sys.argv[4])-1
title = sys.argv[5]

# 参数示例：python sw_combine.py 前端开发-JD.xlsx 前端开发.xlsx 1 1 前端开发

def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name(title)

    sheet_value = []
    for row in sheet.rows:
        tmp = []
        for cell in range(len(row)):
            tmp.append(row[cell].value)
        sheet_value.append(tmp)
    return sheet_value

def write07Excel(path, value):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = title
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    wb.save(path)
    print("写入数据成功！")


def mymovefile(srcfile, dstfile):
    if not os.path.isfile(srcfile):
        print "%s not exist!" % (srcfile)
    else:
        fpath, fname = os.path.split(dstfile)  # 分离文件名和路径
        if not os.path.exists(fpath):
            os.makedirs(fpath)  # 创建路径
        shutil.move(srcfile, dstfile)  # 移动文件
        print "move %s -> %s" % (srcfile, dstfile)


def mycopyfile(srcfile, dstfile):
    if not os.path.isfile(srcfile):
        print "%s not exist!" % (srcfile)
        return False
    else:
        fpath, fname = os.path.split(dstfile)  # 分离文件名和路径
        if not os.path.exists(fpath):
            os.makedirs(fpath)  # 创建路径
        shutil.copyfile(srcfile, dstfile)  # 复制文件
        print "copy %s -> %s" % (srcfile, dstfile)
        return True


srcfile = Pos_Table2
dstfile = './data/combine/' + sys.argv[2]

Table1_txt = read07Excel(Pos_Table1)
Table2_txt = read07Excel(Pos_Table2)

total = []
for i in range(len(Table1_txt)):
    tmp = []
    for j in range(len(Table2_txt)):
        if str(Table1_txt[i][Count1]) == str(Table2_txt[j][Count2]):
            for x in range(len(Table1_txt[i])):
                if x == Count1:
                    continue
                Table2_txt[j].append(Table1_txt[i][x])
            tmp = Table2_txt[j]
            total.append(tmp)
            break

if mycopyfile(srcfile, dstfile) == True:
    write07Excel(Pos_Table_End, total)
else:
    print("写入数据失败！")
