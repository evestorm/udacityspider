import xlrd
import xlwt
import openpyxl
import string
import sys
import os
import shutil
import pymysql


JOB_LIST_DIR = './data/list/'
JOB_DETAIL_DIR = './data/detail/'

def read07Excel(path, title):
    wb = openpyxl.load_workbook(path)
    sheet = wb[title]

    sheet_value = []
    for row in sheet.rows:
        tmp = []
        for cell in range(len(row)):
            tmp.append(row[cell].value)
        sheet_value.append(tmp)
    return sheet_value

# path：目标路径，value：total
def write07Excel(path, value, title):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = title
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    wb.save(path)
    print("写入数据成功！")


def mymovefile(srcfile, dstfile):
    if not os.path.isfile(srcfile):
        print("%s not exist!" % (srcfile))
    else:
        fpath, fname = os.path.split(dstfile)  # 分离文件名和路径
        if not os.path.exists(fpath):
            os.makedirs(fpath)  # 创建路径
        shutil.move(srcfile, dstfile)  # 移动文件
        print("move %s -> %s" % (srcfile, dstfile))


def mycopyfile(srcfile, dstfile):
    if not os.path.isfile(srcfile):
        print("%s not exist!" % (srcfile))
        return False
    else:
        fpath, fname = os.path.split(dstfile)  # 分离文件名和路径
        if not os.path.exists(fpath):
            os.makedirs(fpath)  # 创建路径
        shutil.copyfile(srcfile, dstfile)  # 复制文件
        print("copy %s -> %s" % (srcfile, dstfile))
        return True


def total_sw(positionName, Count1, Count2):
    Pos_Table1 = JOB_DETAIL_DIR + positionName + '-JD.xlsx'
    Pos_Table2 = JOB_LIST_DIR + positionName + '.xlsx'
    print(Pos_Table1, Pos_Table2)
    Table1_txt = read07Excel(Pos_Table1, positionName)
    Table2_txt = read07Excel(Pos_Table2, positionName)
    total = []
    for i in range(len(Table1_txt)):
        tmp = []
        for j in range(len(Table2_txt)):
            if str(Table1_txt[i][Count1]) == str(Table2_txt[j][Count2]):
                for x in range(len(Table1_txt[i])):
                    if x == Count1:
                        continue
                    Table2_txt[j].append(Table1_txt[i][x])
                    # print(Table2_txt[j])
                tmp = Table2_txt[j]
                total.append(tmp)
                # print(total)
                break

    dstfile = './data/combine/' + positionName + '.xlsx'
    if mycopyfile(Pos_Table2, dstfile) == True:
        write07Excel(dstfile, total, positionName)
    else:
        print("写入数据失败！")

def combine_sw():
    for excel_file in os.listdir(JOB_LIST_DIR):
        positionName = excel_file.replace('.xlsx', '')
        print(positionName)
        total_sw(positionName, 0, 0)


# 连接数据库
db = pymysql.connect(host='127.0.0.1', user='root', password='root', port=8889, db='lagou_job', charset='utf8mb4')
# 使用：add_Mysql(count, job_title, job_salary, job_city, job_experience, job_education, company_name, company_type, company_status, company_people, job_tips, job_welfare)
def add_Mysql(id, job_title, job_salary, job_city, job_experience, job_education, company_name, company_type, company_status, company_people, job_tips, job_welfare):
    # 将数据写入数据库中
    try:
        cursor = db.cursor()
        sql = 'insert into job(id, job_title, job_salary, job_city, job_experience, job_education, company_name, company_type, company_status, company_people, job_tips, job_welfare) values ("%d", "%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s", "%s")' % (
            id, job_title, job_salary, job_city, job_experience, job_education, company_name, company_type, company_status, company_people, job_tips, job_welfare)
        print(sql)
        cursor.execute(sql)
        print(cursor.lastrowid)
        db.commit()
    except Exception as e:
        print(e)
        db.rollback()

combine_sw()
